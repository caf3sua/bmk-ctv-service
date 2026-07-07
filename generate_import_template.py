"""Generate the static Excel template used for bulk checklist import
(templates/collaborator_checklist_template.xlsx). Re-run this whenever the
import column structure changes.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "templates")
TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "collaborator_checklist_template.xlsx")

HEADERS = ["Mã nhân viên", "Họ tên", "CCCD", "Cam kết thuế", "CV", "Thông tin cư trú", "Bằng cấp"]
EXAMPLE_ROW = ["CTV001", "Nguyễn Văn A", "Đã nộp", "", "Đã nộp", "Đã nộp", ""]

def generate():
    os.makedirs(TEMPLATE_DIR, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="004F9F", end_color="004F9F", fill_type="solid")
    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.append(EXAMPLE_ROW)
    for col_idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(16, len(HEADERS[col_idx - 1]) + 6)

    ws.freeze_panes = "A2"

    note_row = 4
    ws.cell(row=note_row, column=1, value="Ghi chú:").font = Font(bold=True, italic=True)
    ws.cell(
        row=note_row + 1,
        column=1,
        value=(
            "- Mã nhân viên phải khớp với mã đã tồn tại trong hệ thống, các mã không tìm thấy sẽ bị bỏ qua."
        ),
    ).font = Font(italic=True, size=10)
    ws.cell(
        row=note_row + 2,
        column=1,
        value="- Cột CCCD, Cam kết thuế, CV, Thông tin cư trú, Bằng cấp: để trống = Chưa nộp, nhập bất kỳ giá trị nào (VD: Đã nộp) = Đã nộp.",
    ).font = Font(italic=True, size=10)

    wb.save(TEMPLATE_PATH)
    print(f"Generated template at {TEMPLATE_PATH}")

if __name__ == "__main__":
    generate()
