import os
from datetime import date, datetime, timezone
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
from typing import List
from app.core.activity_log import record_activity
from app.core.database import get_db
from app.core.logging import get_logger
from app.core.security import get_current_user
from app.models.collaborator import CollaboratorCreate, CollaboratorUpdate, CollaboratorResponse
from app.core.s3 import upload_to_s3
from app.core.config import settings


router = APIRouter(prefix="/api/collaborators", tags=["Collaborators"])
logger = get_logger(__name__)

COLLECTION = "bmk_ctv_collaborators"

SERVICE_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMPLATE_PATH = os.path.join(SERVICE_ROOT, "templates", "collaborator_checklist_template.xlsx")

# Template có 3 dòng tiêu đề (row 2-4, do merge cells) rồi mới tới dữ liệu (row 5 trở đi).
HEADER_START_ROW = 2
DATA_START_ROW = 5

EMPLOYEE_CODE_LABEL = "Mã nhân viên"
START_DATE_LABEL = "Ngày bắt đầu"
END_DATE_LABEL = "Ngày kết thúc"
LIQUIDATION_DATE_LABEL = "Biên bản thanh lí"

# (Tên cột trong file Excel, tên field boolean tương ứng trong checklist)
CHECKLIST_COLUMNS = [
    ("CCCD", "submittedIdCard"),
    ("Cam kết thuế", "submittedTaxCommitment"),
    ("CV", "submittedCV"),
    ("Thông tin cư trú", "submittedResidenceInfo"),
    ("Bằng cấp", "submittedDegree"),
]

# (Tên cột trong file Excel, tên field text tương ứng ở hồ sơ CTV) - dùng khi cần tạo mới CTV từ file import
PROFILE_TEXT_COLUMNS = [
    ("Họ tên", "fullName"),
    ("Mã số thuế", "taxCode"),
    ("Số CCCD", "idNumber"),
    ("Email", "email"),
    ("Số điện thoại", "phone"),
    ("Địa chỉ", "address"),
]
DOB_LABEL = "Ngày sinh"

DATE_TEXT_FORMATS = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%y", "%d-%m-%y"]

_UNSET = object()

def _now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

def _parse_excel_date(value) -> str | None:
    """Chuyển giá trị 1 ô Excel (kiểu Date hoặc Text) thành chuỗi ISO 'YYYY-MM-DD'."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return from_excel(value).date().isoformat()

    text = str(value).strip()
    if not text or text in ("-", "—"):
        return None
    for fmt in DATE_TEXT_FORMATS:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"không nhận dạng được định dạng ngày '{text}'")

def _cell_str(row, idx) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()

def _to_response(doc: dict) -> dict:
    doc = dict(doc)
    doc["employeeCode"] = doc["_id"]
    return doc

def _actor_name(current_user: dict) -> str:
    return current_user.get("name") or current_user.get("username", "")

@router.get("", response_model=List[CollaboratorResponse])
async def list_collaborators(db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Fetch all collaborators, sorted by employee code."""
    items = []
    cursor = db[COLLECTION].find({}).sort("_id", 1)
    async for doc in cursor:
        items.append(_to_response(doc))
    return items

@router.get("/template")
async def download_import_template(current_user: dict = Depends(get_current_user)):
    """Download the Excel template used for bulk checklist import."""
    if not os.path.exists(TEMPLATE_PATH):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy file mẫu")
    return FileResponse(
        TEMPLATE_PATH,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="mau_import_checklist_ctv.xlsx",
    )

@router.get("/export")
async def export_collaborators(db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Export all collaborators' checklist status to an Excel file using the template."""
    full_name = _actor_name(current_user)
    if not os.path.exists(TEMPLATE_PATH):
        await record_activity(
            db, action="export_collaborators", result="fail", full_name=full_name,
            username=current_user.get("username", ""),
            message=f"{full_name} đã xuất thất bại danh sách cộng tác viên",
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Không tìm thấy file template mẫu"
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    data_font = Font(name='Aptos Narrow', size=11)

    cursor = db[COLLECTION].find({}).sort("_id", 1)
    row_idx = 5
    stt = 1

    async for doc in cursor:
        checklist = doc.get("checklist")
        if not isinstance(checklist, dict):
            checklist = {}

        contracts = checklist.get("serviceContracts", [])
        start_date = ""
        end_date = ""
        if contracts:
            last_contract = contracts[-1]
            if isinstance(last_contract, dict):
                start_date = last_contract.get("startDate") or ""
                end_date = last_contract.get("endDate") or ""
            else:
                start_date = getattr(last_contract, "startDate", "") or ""
                end_date = getattr(last_contract, "endDate", "") or ""

        row_values = [
            stt,
            doc.get("_id") or doc.get("employeeCode", ""),
            doc.get("fullName", ""),
            doc.get("taxCode", ""),
            doc.get("dob", ""),
            doc.get("idNumber", ""),
            doc.get("email", ""),
            doc.get("phone", ""),
            doc.get("address", ""),
            start_date,
            end_date,
            "Đã nộp" if checklist.get("submittedIdCard") else "",
            "Đã nộp" if checklist.get("submittedTaxCommitment") else "",
            checklist.get("liquidationDate") or "",
            "Đã nộp" if checklist.get("submittedCV") else "",
            "Đã nộp" if checklist.get("submittedResidenceInfo") else "",
            "Đã nộp" if checklist.get("submittedDegree") else "",
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border

            # Alignment formatting
            if col_idx in [1, 2, 4, 5, 6, 8, 10, 11, 12, 13, 14, 15, 16, 17]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_idx += 1
        stt += 1

    # Clear remaining rows in the template if they are beyond our data rows
    for r in range(row_idx, ws.max_row + 1):
        for c in range(1, 18):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.border = Border()

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    exported_count = stt - 1
    await record_activity(
        db, action="export_collaborators", result="success", full_name=full_name,
        username=current_user.get("username", ""),
        message=f"{full_name} đã xuất thành công {exported_count} cộng tác viên",
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=danh_sach_ctv.xlsx"},
    )

@router.post("/import")
async def import_collaborators(
    file: UploadFile = File(...), db=Depends(get_db), current_user: dict = Depends(get_current_user)
):
    """Bulk-create/update collaborators from an uploaded Excel file (see /template) and log to S3/MongoDB history."""
    username = current_user.get("username", "unknown")
    full_name = _actor_name(current_user)
    logger.info(f"Bắt đầu import '{file.filename}' bởi user='{username}'")

    # Generate unique S3 Key and read contents
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    s3_key = f"excel/{timestamp}_{file.filename}"
    s3_bucket = settings.S3_BUCKET

    content = await file.read()

    # 1. Upload to S3
    try:
        upload_to_s3(content, s3_key)
    except Exception as s3_err:
        error_msg = f"Lỗi upload S3: {str(s3_err)}"
        await db["bmk_ctv_upload_history"].insert_one({
            "filename": file.filename,
            "s3Key": s3_key,
            "s3Bucket": s3_bucket,
            "uploadedBy": full_name,
            "username": username,
            "rowsProcessed": 0,
            "createdCount": 0,
            "updatedCount": 0,
            "status": "fail",
            "message": error_msg,
            "createdAt": _now()
        })
        await record_activity(
            db, action="import_collaborators", result="fail", full_name=full_name, username=username,
            message=f"{full_name} đã nhập thất bại 0 cộng tác viên ({error_msg})",
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Không thể lưu trữ file lên S3: {str(s3_err)}"
        )

    # 2. Process logic
    try:
        if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Chỉ hỗ trợ file Excel (.xlsx)")

        try:
            wb = load_workbook(BytesIO(content), data_only=True)
        except Exception:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Không đọc được file Excel, vui lòng dùng đúng file mẫu")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < DATA_START_ROW:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File không có dữ liệu")

        header_row_slices = rows[HEADER_START_ROW - 1 : DATA_START_ROW - 1]
        header = [
            next((str(cell).strip() for cell in reversed(col_cells) if cell not in (None, "")), "")
            for col_cells in zip(*header_row_slices)
        ]

        if EMPLOYEE_CODE_LABEL not in header:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Không tìm thấy cột 'Mã nhân viên' trong file, vui lòng dùng đúng file mẫu",
            )
        code_idx = header.index(EMPLOYEE_CODE_LABEL)
        start_date_idx = header.index(START_DATE_LABEL) if START_DATE_LABEL in header else None
        end_date_idx = header.index(END_DATE_LABEL) if END_DATE_LABEL in header else None
        liquidation_idx = header.index(LIQUIDATION_DATE_LABEL) if LIQUIDATION_DATE_LABEL in header else None

        column_indices = {field: header.index(label) for label, field in CHECKLIST_COLUMNS if label in header}
        profile_indices = {field: header.index(label) for label, field in PROFILE_TEXT_COLUMNS if label in header}
        dob_idx = header.index(DOB_LABEL) if DOB_LABEL in header else None

        updated: List[str] = []
        created: List[str] = []
        date_errors: List[str] = []

        for current_row_num, row in enumerate(rows[DATA_START_ROW - 1 :], start=DATA_START_ROW):
            if row is None or all(cell in (None, "") for cell in row):
                continue

            raw_code = row[code_idx] if code_idx < len(row) else None
            employee_code = str(raw_code).strip() if raw_code is not None else ""
            if not employee_code:
                if not all(cell in (None, "") for cell in row):
                    date_errors.append(f"Dòng {current_row_num}: Thiếu Mã nhân viên")
                continue

            existing = await db[COLLECTION].find_one({"_id": employee_code})

            checklist_values = {}
            for field, col_idx in column_indices.items():
                value = row[col_idx] if col_idx < len(row) else None
                checklist_values[field] = value is not None and str(value).strip() != ""

            def parse_cell_date(idx):
                if idx is None or idx >= len(row):
                    return _UNSET
                raw = row[idx]
                if raw is None or (isinstance(raw, str) and not raw.strip()):
                    return None
                try:
                    return _parse_excel_date(raw)
                except ValueError as exc:
                    date_errors.append(f"Dòng {current_row_num} (Mã NV: {employee_code}), cột '{header[idx]}': {exc}")
                    return _UNSET

            start_val = parse_cell_date(start_date_idx)
            end_val = parse_cell_date(end_date_idx)
            liquidation_val = parse_cell_date(liquidation_idx)

            if existing:
                updates = {f"checklist.{field}": val for field, val in checklist_values.items()}

                if start_date_idx is not None or end_date_idx is not None:
                    contracts = existing.get("checklist", {}).get("serviceContracts") or []
                    last_contract = dict(contracts[-1]) if contracts and isinstance(contracts[-1], dict) else {}
                    if start_val is not _UNSET:
                        last_contract["startDate"] = start_val
                    if end_val is not _UNSET:
                        last_contract["endDate"] = end_val
                    updates["checklist.serviceContracts"] = (contracts[:-1] if contracts else []) + [last_contract]

                if liquidation_val is not _UNSET:
                    updates["checklist.liquidationDate"] = liquidation_val

                updates["updatedAt"] = _now()
                await db[COLLECTION].update_one({"_id": employee_code}, {"$set": updates})
                updated.append(employee_code)
            else:
                now_ts = _now()
                dob_val = parse_cell_date(dob_idx)
                new_doc = {
                    "_id": employee_code,
                    "employeeCode": employee_code,
                    "fullName": _cell_str(row, profile_indices.get("fullName")),
                    "taxCode": _cell_str(row, profile_indices.get("taxCode")),
                    "dob": None if dob_val is _UNSET else dob_val,
                    "idNumber": _cell_str(row, profile_indices.get("idNumber")),
                    "email": _cell_str(row, profile_indices.get("email")),
                    "phone": _cell_str(row, profile_indices.get("phone")),
                    "address": _cell_str(row, profile_indices.get("address")),
                    "checklist": {
                        **checklist_values,
                        "serviceContracts": [{
                            "startDate": None if start_val is _UNSET else start_val,
                            "endDate": None if end_val is _UNSET else end_val,
                        }],
                        "liquidationDate": None if liquidation_val is _UNSET else liquidation_val,
                    },
                    "createdAt": now_ts,
                    "updatedAt": now_ts,
                }
                await db[COLLECTION].insert_one(new_doc)
                created.append(employee_code)

        total_processed = len(created) + len(updated)
        logger.info(
            f"Import '{file.filename}' hoàn tất bởi user='{username}': "
            f"tạo mới={len(created)}, cập nhật={len(updated)}, lỗi ngày tháng={len(date_errors)}"
        )
        if date_errors:
            logger.warning(f"Import '{file.filename}' có {len(date_errors)} lỗi định dạng ngày: {date_errors}")

        success_msg = f"Nhập thành công {total_processed} cộng tác viên (Tạo mới: {len(created)}, Cập nhật: {len(updated)})"
        if date_errors:
            success_msg += f". Có {len(date_errors)} lỗi dữ liệu:\n" + "\n".join(f"- {err}" for err in date_errors)

        await record_activity(
            db, action="import_collaborators", result="success", full_name=full_name, username=username,
            message=f"{full_name} đã nhập thành công {total_processed} cộng tác viên",
        )

        # Record success history
        await db["bmk_ctv_upload_history"].insert_one({
            "filename": file.filename,
            "s3Key": s3_key,
            "s3Bucket": s3_bucket,
            "uploadedBy": full_name,
            "username": username,
            "rowsProcessed": total_processed,
            "createdCount": len(created),
            "updatedCount": len(updated),
            "status": "success",
            "message": success_msg,
            "createdAt": _now()
        })

        return {
            "updatedCount": len(updated),
            "updated": updated,
            "createdCount": len(created),
            "created": created,
            "dateErrorCount": len(date_errors),
            "dateErrors": date_errors,
        }

    except HTTPException as he:
        logger.warning(f"Import thất bại (HTTP {he.status_code}): {he.detail}")
        await db["bmk_ctv_upload_history"].insert_one({
            "filename": file.filename,
            "s3Key": s3_key,
            "s3Bucket": s3_bucket,
            "uploadedBy": full_name,
            "username": username,
            "rowsProcessed": 0,
            "createdCount": 0,
            "updatedCount": 0,
            "status": "fail",
            "message": he.detail,
            "createdAt": _now()
        })
        await record_activity(
            db, action="import_collaborators", result="fail", full_name=full_name, username=username,
            message=f"{full_name} đã nhập thất bại 0 cộng tác viên: {he.detail}",
        )
        raise he

    except Exception as exc:
        error_detail = str(exc)
        logger.error(f"Import thất bại (Lỗi không xác định): {error_detail}")
        await db["bmk_ctv_upload_history"].insert_one({
            "filename": file.filename,
            "s3Key": s3_key,
            "s3Bucket": s3_bucket,
            "uploadedBy": full_name,
            "username": username,
            "rowsProcessed": 0,
            "createdCount": 0,
            "updatedCount": 0,
            "status": "fail",
            "message": f"Lỗi hệ thống: {error_detail}",
            "createdAt": _now()
        })
        await record_activity(
            db, action="import_collaborators", result="fail", full_name=full_name, username=username,
            message=f"{full_name} đã nhập thất bại 0 cộng tác viên: {error_detail}",
        )
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Không đọc được file hoặc lỗi xử lý dữ liệu: {error_detail}")

@router.get("/{employee_code}", response_model=CollaboratorResponse)
async def get_collaborator(employee_code: str, db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Fetch a single collaborator by employee code."""
    doc = await db[COLLECTION].find_one({"_id": employee_code})
    if not doc:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Không tìm thấy cộng tác viên "{employee_code}"'
        )
    return _to_response(doc)

@router.post("", response_model=CollaboratorResponse, status_code=status.HTTP_201_CREATED)
async def create_collaborator(payload: CollaboratorCreate, db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Create a new collaborator profile."""
    full_name = _actor_name(current_user)
    employee_code = payload.employeeCode.strip()
    if not employee_code:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Mã nhân viên là bắt buộc")

    existing = await db[COLLECTION].find_one({"_id": employee_code})
    if existing:
        await record_activity(
            db, action="create_collaborator", result="fail", full_name=full_name,
            username=current_user.get("username", ""),
            message=f"{full_name} tạo thất bại hồ sơ cho cộng tác viên mã {employee_code}",
        )
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f'Mã nhân viên "{employee_code}" đã tồn tại'
        )

    now = _now()
    doc = payload.model_dump()
    doc["employeeCode"] = employee_code
    doc["_id"] = employee_code
    doc["createdAt"] = now
    doc["updatedAt"] = now
    await db[COLLECTION].insert_one(doc)
    await record_activity(
        db, action="create_collaborator", result="success", full_name=full_name,
        username=current_user.get("username", ""),
        message=f"{full_name} tạo thành công hồ sơ cho cộng tác viên mã {employee_code}",
    )
    return _to_response(doc)

@router.put("/{employee_code}", response_model=CollaboratorResponse)
async def update_collaborator(employee_code: str, payload: CollaboratorUpdate, db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Update an existing collaborator profile."""
    full_name = _actor_name(current_user)
    username = current_user.get("username", "")

    existing = await db[COLLECTION].find_one({"_id": employee_code})
    if not existing:
        await record_activity(
            db, action="update_collaborator", result="fail", full_name=full_name, username=username,
            message=f"{full_name} cập nhật thất bại hồ sơ cho cộng tác viên mã {employee_code}",
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Không tìm thấy cộng tác viên "{employee_code}"'
        )

    new_code = payload.employeeCode.strip() or employee_code
    if new_code != employee_code:
        conflict = await db[COLLECTION].find_one({"_id": new_code})
        if conflict:
            await record_activity(
                db, action="update_collaborator", result="fail", full_name=full_name, username=username,
                message=f"{full_name} cập nhật thất bại hồ sơ cho cộng tác viên mã {employee_code}",
            )
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f'Mã nhân viên "{new_code}" đã tồn tại'
            )

    doc = payload.model_dump()
    doc["employeeCode"] = new_code
    doc["_id"] = new_code
    doc["createdAt"] = existing["createdAt"]
    doc["updatedAt"] = _now()

    if new_code != employee_code:
        await db[COLLECTION].delete_one({"_id": employee_code})
    await db[COLLECTION].replace_one({"_id": new_code}, doc, upsert=True)
    await record_activity(
        db, action="update_collaborator", result="success", full_name=full_name, username=username,
        message=f"{full_name} cập nhật thành công hồ sơ cho cộng tác viên mã {new_code}",
    )
    return _to_response(doc)

@router.delete("/{employee_code}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_collaborator(employee_code: str, db=Depends(get_db), current_user: dict = Depends(get_current_user)):
    """Delete a collaborator profile by employee code."""
    full_name = _actor_name(current_user)
    username = current_user.get("username", "")

    result = await db[COLLECTION].delete_one({"_id": employee_code})
    if result.deleted_count == 0:
        await record_activity(
            db, action="delete_collaborator", result="fail", full_name=full_name, username=username,
            message=f"{full_name} xóa thất bại hồ sơ cộng tác viên mã {employee_code}",
        )
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f'Không tìm thấy cộng tác viên "{employee_code}"'
        )

    await record_activity(
        db, action="delete_collaborator", result="success", full_name=full_name, username=username,
        message=f"{full_name} xóa thành công hồ sơ cộng tác viên mã {employee_code}",
    )
    return None
